#!/usr/bin/env python
# 批量转换采样率,语音全部需要提前转换采样率才可以使用

import os
import librosa
import soundfile
import json
import random

import openpyxl
import openpyxl
from openpyxl import load_workbook


# 定义转换采样率的函数，接收3个变量：原音频路径、重新采样后的音频存储路径、目标采样率
# 定义转换采样率的函数，接收3个变量：原音频路径、重新采样后的音频存储路径、目标采样率
def change_sample_rate(path, new_dir_path, new_sample_rate):
    wavfile = path.split('/')[-1]  # 提取音频文件名，如“1.wav"
    # new_file_name = wavfile.split('.')[0] + '_8k.wav'      #此行代码可用于对转换后的文件进行重命名（如有需要）

    signal, sr = librosa.load(path, sr=None)  # 调用librosa载入音频
    try:
        new_signal = librosa.resample(signal, sr, new_sample_rate)  # 调用librosa进行音频采样率转换
    except Exception as e:
        pass

    new_path = os.path.join(new_dir_path, wavfile)  # 指定输出音频的路径，音频文件与原音频同名
    # new_path = os.path.join(new_dir_path, new_file_name)      #若需要改名则启用此行代码
    print(new_path)

    # librosa.output.write_wav(new_path, new_signal , new_sample_rate)      #因版本问题，此方法可能用不了
    soundfile.write(new_path, new_signal, new_sample_rate)


def getInfo():
    import os
    for ia in os.listdir():
        if ia.endswith('.mp4') or ia.endswith('.mp3') or ia.endswith('.wav'):
            video=ia
    os.system("ffmpeg -i "+video+" -f wav 123.wav")
    for ia in os.listdir():
        print("-----"+ia)
        if ia.endswith('.xlsx') and ia!='dataSet.xlsx':
            xlsxPath =ia
        if ia.endswith('.wav'):
            audio_in_dir =ia
    # 第一步打开工作簿
    wb = openpyxl.load_workbook(xlsxPath)
    # 第二步选取表单
    sheet = wb.active
    # 按行获取数据转换成列表
    rows_data = list(sheet.rows)
    # 获取表单的表头信息(第一行)，也就是列表的第一个元素
    titles = [title.value for title in rows_data.pop(0)]
    print(titles)


    all_row_dict = []


    #新建词库，当你需要新建词库时取消注释
    newDict={}

    # 遍历出除了第一行的其他行
    for a_row in rows_data:
        the_row_data = [cell.value for cell in a_row]
        # 将表头和该条数据内容，打包成一个字典
        row_dict = dict(zip(titles, the_row_data))
        all_row_dict.append(row_dict)

    wba = openpyxl.load_workbook("dataSet.xlsx")
    sheeta = wba.active
    for i in all_row_dict:

        startTime=i.get(titles[0])
        endTime=i.get(titles[1])
        text = i.get(titles[2])
        print(startTime+'  '+endTime+'  '+text)

        random_str = ''
        base_str = 'ABCDEFGHIGKLMNOPQRSTUVWXYZabcdefghigklmnopqrstuvwxyz0123456789'
        for i in range(7):
            random_str += base_str[random.randint(0, len(base_str)-1)]

        audio_out_dir = "wavss\\"+random_str+".wav"

        audio_cut(audio_in_dir,audio_out_dir,startTime,endTime)
        sheeta.append([random_str+ '.wav',text]) 
        #addSub([random_str+ '.wav',text])
        igo = 'wavs/'+random_str+ '.wav|' + text+'\n'
        print('wavs/'+random_str+ '.wav|' + text)



        file = open('filelists/list.txt', 'a')
        file.write(igo)
        file.close()
    wba.save("dataSet.xlsx")

def audio_cut(audio_in_path, audio_out_path, startTime, endTime):
    """
    :param audio_in_path: 输入音频的绝对路径
    :param audio_out_path: 切分后输出音频的绝对路径
    :param start_time: 切分开始时间
    :param dur_time: 切分持续时间
    :return:
    """
    os.system("ffmpeg  -i "+audio_in_path+"  -vn -acodec copy -ss "+startTime+" -to "+endTime+" "+audio_out_path)
    #os.system("ffmpeg  -i source.mp3  -vn -acodec copy -ss 00:03:21.36 -t 00:00:41 output.mp3")
def addSub(c):
    wba = openpyxl.load_workbook("dataSet.xlsx")
    sheeta = wba.active
    sheeta.append(c)  # 插入一行数据
    wba.save("dataSet.xlsx")  # 保存,传入原文件则在
def clearSheet():
    i=0

    while i<4:

        wbs = load_workbook("sub.xlsx")
        wss = wbs.active
        wss.delete_cols(1)  # 删除第 1 列数据
        wbs.save("sub.xlsx")
        i+=1
        print('clear')
if __name__ == '__main__':
    import sys

    print(sys.path)
    print(sys.executable)
    getInfo()

    print('已完成，即将执行音频采样率转换')
    # 指定原音频文件夹路径
    original_path = "wavss/"
    wav_list = os.listdir(original_path)

    # 指定转换后的音频文件夹路径
    new_dir_path = "wavs/"
    os.makedirs(new_dir_path, exist_ok=True)

    # 开始以对原音频文件夹内的音频进行采样率的批量转换
    for i in wav_list:
        wav_path = os.path.join(original_path, i)
        change_sample_rate(wav_path, new_dir_path, new_sample_rate=22050)
    import codecs
    import os


    f = codecs.open("filelists/list.txt", 'r', 'ansi')
    ff = f.read()
    file_object = codecs.open('filelists/list.txt', 'w', 'utf-8')
    file_object.write(ff)
    print('已完成，语音文件在 wavs/ 文件夹下\n标注信息在filelists/list.txt')